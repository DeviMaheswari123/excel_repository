import openpyxl
from random import randint

# Load the workbook
wb = openpyxl.Workbook()

# Access the default worksheet
ws = wb.active

# Generate random numbers
num1 = randint(1, 10)
num2 = randint(1, 10)

# Perform the calculation
result = num1 + num2

# Write the numbers and result to the worksheet
ws['A1'] = num1
ws['B1'] = '+'
ws['C1'] = num2
ws['D1'] = '='
ws['E1'] = result

# Save the workbook
wb.save("random_calculation2.xlsx")